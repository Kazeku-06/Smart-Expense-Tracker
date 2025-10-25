import io
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from models import Transaction, Category
from flask_jwt_extended import get_jwt_identity

def generate_pdf_report(start_date=None, end_date=None, user_id=None):
    """Generate PDF report untuk transaksi user"""
    try:
        # Query transactions dengan filter tanggal
        query = Transaction.query.filter_by(user_id=user_id)
        
        if start_date:
            query = query.filter(Transaction.date >= start_date)
        if end_date:
            query = query.filter(Transaction.date <= end_date)
            
        transactions = query.order_by(Transaction.date.desc()).all()
        
        # Buat buffer untuk PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Judul laporan
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center alignment
        title = Paragraph("Laporan Pengeluaran Pribadi", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Informasi periode
        period_text = f"Periode: {start_date.strftime('%d/%m/%Y') if start_date else 'Semua Waktu'} - {end_date.strftime('%d/%m/%Y') if end_date else 'Sekarang'}"
        period_style = styles['Normal']
        period_style.alignment = 1
        period = Paragraph(period_text, period_style)
        elements.append(period)
        elements.append(Spacer(1, 20))
        
        # Data transaksi
        if transactions:
            # Header table
            data = [['Tanggal', 'Kategori', 'Deskripsi', 'Jumlah (Rp)']]
            
            # Data transaksi
            total_amount = 0
            for transaction in transactions:
                date_str = transaction.date.strftime('%d/%m/%Y')
                category_name = transaction.category.name
                description = transaction.description
                amount = f"{transaction.amount:,.0f}"
                total_amount += transaction.amount
                
                data.append([date_str, category_name, description, amount])
            
            # Total row
            data.append(['', '', 'TOTAL', f"{total_amount:,.0f}"])
            
            # Buat table
            table = Table(data, colWidths=[80, 100, 200, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
            
            # Summary
            summary_text = f"Total Transaksi: {len(transactions)} | Total Pengeluaran: Rp {total_amount:,.0f}"
            summary = Paragraph(summary_text, styles['Heading2'])
            elements.append(summary)
        else:
            # Tidak ada data
            no_data = Paragraph("Tidak ada data transaksi untuk periode yang dipilih.", styles['BodyText'])
            elements.append(no_data)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer, f"laporan_pengeluaran_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
    except Exception as e:
        raise Exception(f"Error generating PDF: {str(e)}")

def generate_excel_report(start_date=None, end_date=None, user_id=None):
    """Generate Excel report untuk transaksi user"""
    try:
        # Query transactions dengan filter tanggal
        query = Transaction.query.filter_by(user_id=user_id)
        
        if start_date:
            query = query.filter(Transaction.date >= start_date)
        if end_date:
            query = query.filter(Transaction.date <= end_date)
            
        transactions = query.order_by(Transaction.date.desc()).all()
        
        # Buat workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Pengeluaran"
        
        # Header
        headers = ['Tanggal', 'Kategori', 'Deskripsi', 'Jumlah (Rp)']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data transaksi
        total_amount = 0
        for transaction in transactions:
            date_str = transaction.date.strftime('%d/%m/%Y')
            category_name = transaction.category.name
            description = transaction.description
            amount = transaction.amount
            total_amount += amount
            
            ws.append([date_str, category_name, description, amount])
        
        # Total row
        if transactions:
            ws.append(['', '', 'TOTAL', total_amount])
            
            # Style total row
            total_row = ws[ws.max_row]
            for cell in total_row:
                cell.font = Font(bold=True)
        
        # Auto adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save ke buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer, f"laporan_pengeluaran_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
    except Exception as e:
        raise Exception(f"Error generating Excel: {str(e)}")